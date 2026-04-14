import openpyxl as op
from openpyxl.styles import Alignment, PatternFill
import random

wb = op.Workbook()
ws = wb.active
ws.title = 'Harmonogram'

for i in range(101):
    column = i + 2
    cell = ws.cell(row = 17, column = column, value = i)
    cell.alignment = Alignment(horizontal = 'center')
    ws.column_dimensions[op.utils.cell.get_column_letter(column)].width = 3
    


colors = [
    "FF0000", # Czerwony
    "FFFF00", # Żółty
    "00DBFF", # jasnoniebieski
    "00FF11", # jasnozielony
    "004EFF", # ciemnoniebieski
    "BA4EFF", # fioletowy
    "858B83", # szary
    "008100", # ciemnozielony
    "F88100", # pomarańczowy
]

fills = [PatternFill(start_color = color, end_color= color, fill_type='solid') for color in colors]

# Macierz C
C = [
    [10, 20, 30, 40, 50, 60, 70, 80, 101],
    [11, 21, 31, 41, 51, 61, 71, 81, 101],
    [12, 22, 32, 42, 52, 62, 72, 82, 101],
    [13, 23, 33, 43, 53, 63, 73, 83, 101],
    [14, 24, 34, 44, 54, 64, 74, 84, 101],
    [15, 25, 35, 45, 55, 65, 75, 85, 101],
    [16, 26, 36, 46, 56, 66, 76, 86, 101],
    [17, 27, 37, 47, 57, 67, 77, 87, 101]
]

zadania = list(range(9))
random.shuffle(zadania) 

abstr_plan = []

for i in range(8):
    #przesuniecie zadania o i miejsca, aby się nie nakładały w pionie
    zadania = zadania[-i:] + zadania[:-i] if i > 0 else list(zadania)

    czasy_trwania = []
    czas_poczatkowy = 0

    for czas_koncowy in C[i]:
        czasy_trwania.append(czas_koncowy - czas_poczatkowy)
        czas_poczatkowy = czas_koncowy
    abstr_plan.append({'zadania': zadania, 'czasy_trwania': czasy_trwania})

# randomizacja planu
random.shuffle(abstr_plan)

# Rysowanie w Excelu
for index_maszyny, plan in enumerate(abstr_plan):
    numer_wiersza = index_maszyny * 2 + 2
    ws.cell( row = numer_wiersza, column=1, value=f"Maszyna {index_maszyny + 1}")

    aktualny_czas = 0
    for index_zadania, czas_trwania in zip(plan['zadania'], plan['czasy_trwania']):
        wypelnienie = fills[index_zadania]
        for t in range(czas_trwania):
            numer_kolumny = aktualny_czas + t + 2
            ws.cell(row = numer_wiersza, column= numer_kolumny).fill = wypelnienie
        aktualny_czas += czas_trwania

# Generowanie legendy 
wiersz_legendy = 19
ws.cell(row=wiersz_legendy-1, column=2, value="Zadania:")
for i in range(9):
    ws.cell(row=wiersz_legendy + i, column=2).fill=fills[i]
    ws.cell(row=wiersz_legendy + i, column=3, value=f'Zadanie {i + 1}')

# Zmiana szerokości kolumny A
ws.column_dimensions['A'].width = 10
    
wb.save("harmonogram.xlsx")